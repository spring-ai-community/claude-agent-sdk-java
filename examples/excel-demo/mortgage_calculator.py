"""
Mortgage Payment Calculator with Amortization Schedule
Generates an Excel file with loan details and payment breakdown
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

def create_mortgage_calculator():
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Mortgage Calculator"

    # Define styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    input_fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
    result_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
    currency_format = '"$"#,##0.00'
    percent_format = '0.00%'

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # ============ LOAN INPUTS SECTION ============
    ws['B2'] = "MORTGAGE PAYMENT CALCULATOR"
    ws['B2'].font = Font(bold=True, size=18, color="2F5496")
    ws.merge_cells('B2:E2')

    # Input labels and cells
    ws['B4'] = "LOAN INPUTS"
    ws['B4'].font = header_font
    ws['B4'].fill = header_fill
    ws.merge_cells('B4:E4')

    inputs = [
        ("B6", "Loan Amount:", "C6", 300000, currency_format),
        ("B7", "Annual Interest Rate:", "C7", 0.065, percent_format),
        ("B8", "Loan Term (Years):", "C8", 30, "0"),
        ("B9", "Start Date:", "C9", "2025-01-01", "YYYY-MM-DD"),
    ]

    for label_cell, label, value_cell, value, fmt in inputs:
        ws[label_cell] = label
        ws[label_cell].font = Font(bold=True)
        ws[value_cell] = value
        ws[value_cell].fill = input_fill
        ws[value_cell].border = thin_border
        if fmt != "YYYY-MM-DD":
            ws[value_cell].number_format = fmt

    # ============ CALCULATED RESULTS SECTION ============
    ws['B11'] = "CALCULATED RESULTS"
    ws['B11'].font = header_font
    ws['B11'].fill = header_fill
    ws.merge_cells('B11:E11')

    # Monthly Payment Formula using Excel's PMT function
    ws['B13'] = "Monthly Payment:"
    ws['B13'].font = Font(bold=True)
    ws['C13'] = "=-PMT(C7/12, C8*12, C6)"
    ws['C13'].number_format = currency_format
    ws['C13'].fill = result_fill
    ws['C13'].border = thin_border

    ws['B14'] = "Total Payments:"
    ws['B14'].font = Font(bold=True)
    ws['C14'] = "=C8*12"
    ws['C14'].fill = result_fill
    ws['C14'].border = thin_border

    ws['B15'] = "Total Amount Paid:"
    ws['B15'].font = Font(bold=True)
    ws['C15'] = "=C13*C14"
    ws['C15'].number_format = currency_format
    ws['C15'].fill = result_fill
    ws['C15'].border = thin_border

    ws['B16'] = "Total Interest Paid:"
    ws['B16'].font = Font(bold=True)
    ws['C16'] = "=C15-C6"
    ws['C16'].number_format = currency_format
    ws['C16'].fill = result_fill
    ws['C16'].border = thin_border

    ws['B17'] = "Interest to Principal Ratio:"
    ws['B17'].font = Font(bold=True)
    ws['C17'] = "=C16/C6"
    ws['C17'].number_format = percent_format
    ws['C17'].fill = result_fill
    ws['C17'].border = thin_border

    # ============ AMORTIZATION SCHEDULE ============
    ws['B20'] = "AMORTIZATION SCHEDULE"
    ws['B20'].font = header_font
    ws['B20'].fill = header_fill
    ws.merge_cells('B20:H20')

    # Schedule headers
    schedule_headers = ["Payment #", "Payment Date", "Payment", "Principal", "Interest", "Extra Payment", "Balance"]
    for col, header in enumerate(schedule_headers, start=2):
        cell = ws.cell(row=22, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Generate amortization rows (up to 360 months for 30-year loan)
    max_payments = 360

    for i in range(1, max_payments + 1):
        row = 22 + i

        # Payment Number
        ws.cell(row=row, column=2).value = f"=IF({i}<=C$8*12, {i}, \"\")"
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center')

        # Payment Date
        ws.cell(row=row, column=3).value = f'=IF(B{row}<>"", EDATE(C$9, {i}-1), "")'
        ws.cell(row=row, column=3).number_format = "MMM-YYYY"

        # Monthly Payment
        ws.cell(row=row, column=4).value = f'=IF(B{row}<>"", C$13, "")'
        ws.cell(row=row, column=4).number_format = currency_format

        # Principal portion
        if i == 1:
            ws.cell(row=row, column=5).value = f'=IF(B{row}<>"", C$13-(C$6*C$7/12), "")'
        else:
            ws.cell(row=row, column=5).value = f'=IF(B{row}<>"", C$13-(H{row-1}*C$7/12), "")'
        ws.cell(row=row, column=5).number_format = currency_format

        # Interest portion
        if i == 1:
            ws.cell(row=row, column=6).value = f'=IF(B{row}<>"", C$6*C$7/12, "")'
        else:
            ws.cell(row=row, column=6).value = f'=IF(B{row}<>"", H{row-1}*C$7/12, "")'
        ws.cell(row=row, column=6).number_format = currency_format

        # Extra Payment (user can input)
        ws.cell(row=row, column=7).value = 0
        ws.cell(row=row, column=7).number_format = currency_format
        ws.cell(row=row, column=7).fill = input_fill

        # Remaining Balance
        if i == 1:
            ws.cell(row=row, column=8).value = f'=IF(B{row}<>"", MAX(0, C$6-E{row}-G{row}), "")'
        else:
            ws.cell(row=row, column=8).value = f'=IF(B{row}<>"", MAX(0, H{row-1}-E{row}-G{row}), "")'
        ws.cell(row=row, column=8).number_format = currency_format

        # Add borders
        for col in range(2, 9):
            ws.cell(row=row, column=col).border = thin_border

    # ============ SUMMARY BY YEAR ============
    ws_summary = wb.create_sheet("Yearly Summary")

    ws_summary['B2'] = "YEARLY PAYMENT SUMMARY"
    ws_summary['B2'].font = Font(bold=True, size=16, color="2F5496")
    ws_summary.merge_cells('B2:F2')

    summary_headers = ["Year", "Principal Paid", "Interest Paid", "Total Paid", "End Balance"]
    for col, header in enumerate(summary_headers, start=2):
        cell = ws_summary.cell(row=4, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Generate yearly summary formulas
    for year in range(1, 31):
        row = 4 + year
        start_payment = (year - 1) * 12 + 1
        end_payment = year * 12

        # Year number
        ws_summary.cell(row=row, column=2).value = f'=IF({year}<=\'Mortgage Calculator\'!C$8, {year}, "")'
        ws_summary.cell(row=row, column=2).alignment = Alignment(horizontal='center')

        # Principal Paid (sum of principal for that year)
        start_row = 22 + start_payment
        end_row = 22 + end_payment
        ws_summary.cell(row=row, column=3).value = f'=IF(B{row}<>"", SUMPRODUCT((\'Mortgage Calculator\'!B$23:B$382>={start_payment})*(\'Mortgage Calculator\'!B$23:B$382<={end_payment})*(\'Mortgage Calculator\'!E$23:E$382)), "")'
        ws_summary.cell(row=row, column=3).number_format = currency_format

        # Interest Paid
        ws_summary.cell(row=row, column=4).value = f'=IF(B{row}<>"", SUMPRODUCT((\'Mortgage Calculator\'!B$23:B$382>={start_payment})*(\'Mortgage Calculator\'!B$23:B$382<={end_payment})*(\'Mortgage Calculator\'!F$23:F$382)), "")'
        ws_summary.cell(row=row, column=4).number_format = currency_format

        # Total Paid
        ws_summary.cell(row=row, column=5).value = f'=IF(B{row}<>"", C{row}+D{row}, "")'
        ws_summary.cell(row=row, column=5).number_format = currency_format

        # End Balance
        balance_row = 22 + end_payment
        ws_summary.cell(row=row, column=6).value = f'=IF(B{row}<>"", \'Mortgage Calculator\'!H{balance_row}, "")'
        ws_summary.cell(row=row, column=6).number_format = currency_format

        # Borders
        for col in range(2, 7):
            ws_summary.cell(row=row, column=col).border = thin_border

    # ============ INSTRUCTIONS SHEET ============
    ws_help = wb.create_sheet("Instructions")

    instructions = [
        ("B2", "HOW TO USE THIS MORTGAGE CALCULATOR", Font(bold=True, size=16, color="2F5496")),
        ("B4", "1. ENTER YOUR LOAN DETAILS:", Font(bold=True, size=12)),
        ("B5", "   - Loan Amount: The total amount you're borrowing (e.g., $300,000)", None),
        ("B6", "   - Annual Interest Rate: Enter as decimal (e.g., 6.5% = 0.065)", None),
        ("B7", "   - Loan Term: Number of years (typically 15 or 30)", None),
        ("B8", "   - Start Date: When your first payment begins", None),
        ("B10", "2. VIEW YOUR RESULTS:", Font(bold=True, size=12)),
        ("B11", "   - Monthly Payment: Your fixed monthly payment amount", None),
        ("B12", "   - Total Interest: How much interest you'll pay over the loan life", None),
        ("B14", "3. AMORTIZATION SCHEDULE:", Font(bold=True, size=12)),
        ("B15", "   - Shows each monthly payment broken into principal and interest", None),
        ("B16", "   - Extra Payment column: Add extra payments to see how it affects payoff", None),
        ("B18", "4. YEARLY SUMMARY:", Font(bold=True, size=12)),
        ("B19", "   - See the 'Yearly Summary' tab for annual totals", None),
        ("B21", "TIPS:", Font(bold=True, size=12, color="2F5496")),
        ("B22", "   - Blue cells are INPUT cells - you can modify these values", None),
        ("B23", "   - Green cells show CALCULATED results", None),
        ("B24", "   - Add extra payments to pay off your mortgage faster!", None),
    ]

    for cell, text, font in instructions:
        ws_help[cell] = text
        if font:
            ws_help[cell].font = font

    # Set column widths
    for ws_current in [ws, ws_summary, ws_help]:
        ws_current.column_dimensions['A'].width = 3
        ws_current.column_dimensions['B'].width = 25
        ws_current.column_dimensions['C'].width = 18
        ws_current.column_dimensions['D'].width = 15
        ws_current.column_dimensions['E'].width = 15
        ws_current.column_dimensions['F'].width = 15
        ws_current.column_dimensions['G'].width = 15
        ws_current.column_dimensions['H'].width = 18

    # Save workbook
    filename = "Mortgage_Payment_Calculator.xlsx"
    wb.save(filename)
    print(f"Successfully created: {filename}")
    return filename

if __name__ == "__main__":
    create_mortgage_calculator()
